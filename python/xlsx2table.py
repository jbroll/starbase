#!/usr/bin/env python
#

import sys
import re
import os

include starbase.py

program = os.path.splitext(os.path.basename(sys.argv[0]))[0]

if len(sys.argv) <= 1 :
    print """
	usage: %s file sheet range [-data]
    """ % program
    sys.exit(1)


file  = sys.argv[1]

if not os.path.isfile(file) and program != "table2xlsx":
    print program + ": file not found: " + file
    sys.exit(1)


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.cell import column_index_from_string

if os.path.isfile(file) and len(sys.argv) <= 2:
    try:
	xlsx = load_workbook(file)
    except:
	print program + ": cannot open xlsx woorkbook: " + file
	sys.exit(1)

    print "Sheets"
    print "------"
    print "\n".join( xlsx.get_sheet_names())
    sys.exit(1)

if len(sys.argv) <= 2:
    print """
	usage: %s file sheet range [-data]
    """ % program
    sys.exit(1)
    

sheet = sys.argv[2]
cells = sys.argv[3]

if len(sys.argv) == 5 and sys.argv[4] == "-data" :
    header = False
else:
    header = True

if os.path.isfile(file):
    try:
	xlsx = load_workbook(file)
    except:
	print program + ": cannot open xlsx woorkbook: " + file
	sys.exit(1)
else:
    if program == "table2xlsx":
	xlsx     = Workbook()
	ws       = xlsx.active
	ws.title = sheet

    else:
	print program + ": cannot find xlsx woorkbook: " + file
	sys.exit(1)

try:
    ws = xlsx[sheet]
except:
    if program == "table2xlsx":
	ws = xlsx.create_sheet()
	ws.title = sheet
    else:
	print program + ": cannot find sheet: " + sheet
	sys.exit(1)

if program == "xlsx2table":
    cells = cells.split(":")

    if len(cells) == 1 :
	cell0 = ws[cells[0]]

	col = 0
	while cell0.offset(0, col+1).value != None and cell0.offset(0, col+1).value != "":
	    col = col+1

	row = 0
	while cell0.offset(row+1, 0).value != None and cell0.offset(row+1, 0).value != "":
	    row = row+1

	cells.append(cell0.offset(row, col).coordinate)

    else:
	cell0 = ws[cells[0]]
	cell1 = ws[cells[1]]

	col = column_index_from_string(cell1.column)


    print "source	", file, "	", sheet, "	", "\t".join(cells)
    print ""

    if header:
	header = []
	dashes = []
	for row in ws.range(cell0.coordinate + ":" + cell0.offset(0, col).coordinate):
	    for cell in row:
		name = re.sub('[^a-zA-Z0-9_]', "", str(cell.value)).strip()

		header.append(          name)
		dashes.append("-" * len(name))

	print "\t".join(header)
	print "\t".join(dashes)

	cells[0] = cell0.offset(1, 0).coordinate


    for row in ws.range(cells[0] + ":" + cells[1]):
	tab = ""
	for cell in row:
	    print tab, str(cell.value).strip(),
	    tab = "\t"

	print ""

    sys.exit(0)

if program == "table2xlsx":
    try:
	cell0 = ws[cells]
    except:
	print program + ": cannot find upper left corner of table: " + cells

    table = Starbase(sys.stdin)


    if header:
	for ( i, value ) in enumerate(table ** "headline"):
	    ws[cell0.offset(0, i).coordinate] = value

	cell0 = cell0.offset(1, 0)

    for ( j, row ) in enumerate(table):
	for ( i, value ) in enumerate(row):
	    ws[cell0.offset(j, i).coordinate] = value

    xlsx.save(file)

    sys.exit(0)



print "xlsx2table: unknown function"
sys.exit(1)
